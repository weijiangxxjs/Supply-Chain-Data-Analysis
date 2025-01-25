import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
import tkinter as tk
from tkinter import filedialog
import os
from tkinter import ttk
import threading
import time
import tkinter as tk
import tkinter.font as tkFont

# 全局忽略SettingWithCopyWarning警告
warnings.simplefilter("ignore", category=pd.errors.SettingWithCopyWarning)


def proceed():
    global odm_gr_file_name, suppllier_gr_file_name, shipment_file_name, start_time, total_time
    try:

        # 检查是否所有文件都已选择
        if not all([odm_gr_file_name, suppllier_gr_file_name, shipment_file_name]):
            custom_warning("提示", "请选择清洗所需的文件！")
            return  # 直接退出函数，不执行后续数据处理逻辑

        #假设用时60s
        total_time = 90
        # 设置进度条最大值，这里假设你的proceed函数里主要有一些循环等操作，
        # 可以根据实际的循环次数或者步骤数量来合理设置最大值，这里示例设为100
        start_time = time.time()
        progress_bar["maximum"] = 100
        progress_bar["value"] = 0
        progress_bar.place(relx=0.25, rely=0.88, relwidth=0.5, height=30)
        root.update_idletasks()

        # 使用多线程执行长时间运行的操作
        thread = threading.Thread(target=long_running_operation)
        thread.start()
        update_progress()
    except Exception as e:
        print(f"执行 proceed 函数出现错误: {e}")
        # 异常处理，隐藏进度条
        progress_bar.place_forget()
        # 清空文件选择相关显示信息示例（可根据实际需求调整）
        odm_gr_label.config(text="")
        suppllier_gr_label.config(text="")
        shipment_label.config(text="")
        result_label.config(text="数据处理出现异常，请检查文件或联系管理员！")
        result_label.update_idletasks()

#更新进度条函数
def update_progress():
    elapsed_time = time.time() - start_time
    progress_bar["value"] = min(elapsed_time / total_time * 100, 100)
    if progress_bar["value"] < 100:
        root.after(100, update_progress)

#数据清洗
def long_running_operation():
    try:
        # 先显示数据清洗中的提示信息
        result_label.config(text="数据清洗中，请稍等！")
        result_label.update_idletasks()

        # 读取两个 Excel 文件
        # 定义ODM_data的列名列表
        expected_odm_columns = ['BU', 'Month', 'ODM', 'Vendor Name', 'ODM PN', 'Lenovo PN', 'Qty', 'Commodity']
        expected_supplier_columns = ['BU', 'ODM', 'Commodity', 'ODM PN', 'Lenovo PN', '标准命名', 'Vendor Name', 'Qty', 'Month'] 
        # 先处理 ODM_data
        try:
            df_temp = pd.read_excel(odm_gr_file_name, engine='openpyxl', nrows=0)  # 先只读取表头那一行（nrows=0表示只读取列名信息）
            actual_columns = df_temp.columns.tolist()  # 获取实际的列名列表
            if all(col in expected_odm_columns for col in actual_columns):
                ODM_data = pd.read_excel(odm_gr_file_name, engine='openpyxl')
            else:
                print("ODM GR.xlsx 文件的列名不符合要求，请检查文件列名设置！")
                custom_warning("提示", "ODM文件的列名不符合要求，请确认选择的文件是否正确！")
                raise ValueError("列名不符合要求")
                
        except FileNotFoundError:
            print(f"找不到文件 {odm_gr_file_name}，请确认文件路径是否正确。")
        except Exception as e:
            print(f"读取 ODM GR.xlsx 文件出现其他错误: {e}")
            # 异常处理，隐藏进度条
            progress_bar.place_forget()
            # 清空文件选择相关显示信息示例（可根据实际需求调整）
            odm_gr_label.config(text="")
            suppllier_gr_label.config(text="")
            shipment_label.config(text="")
            result_label.config(text="数据处理出现异常，请检查文件或联系管理员！")
            result_label.update_idletasks()

        try:
            df_temp = pd.read_excel(suppllier_gr_file_name, engine='openpyxl', nrows=0)
            actual_columns = df_temp.columns.tolist()
            if all(col in expected_supplier_columns for col in actual_columns):
                Supplier_data = pd.read_excel(suppllier_gr_file_name, engine='openpyxl')
            else:
                print("Supplier GR.xlsx 文件的列名不符合要求，请检查文件列名设置！")
                custom_warning("提示", "Supplier文件的列名不符合要求，请确认选择的文件是否正确！")
                raise ValueError("列名不符合要求")
                
        except FileNotFoundError:
            print(f"找不到文件 {suppllier_gr_file_name}，请确认文件路径是否正确。")
        except Exception as e:
            print(f"读取 Supplier GR.xlsx 文件出现其他错误: {e}")
            # 异常处理，隐藏进度条
            progress_bar.place_forget()
            # 清空文件选择相关显示信息示例（可根据实际需求调整）
            odm_gr_label.config(text="")
            suppllier_gr_label.config(text="")
            shipment_label.config(text="")
            result_label.config(text="数据处理出现异常，请检查文件或联系管理员！")
            result_label.update_idletasks()
        
        
        # 假设要读取的列名为'Column1', 'Column2', 'Column3'（示例，根据实际情况替换为真实列名）
        columns_to_read = ['季度', '物料类型', 'ODM', 'KB Spec', 'Suppliers', 'Share']
        # 先读取表头获取列名信息，判断是否包含指定列
        try:
            df_temp = pd.read_excel(shipment_file_name, sheet_name='Summary', engine='openpyxl', nrows=0)
            actual_columns = df_temp.columns.tolist()
            if all(column in actual_columns for column in columns_to_read):
                TNB_KB_shipment_original = pd.read_excel(shipment_file_name, sheet_name='Summary', engine='openpyxl', usecols=columns_to_read)
            else:
                print(f"文件 {shipment_file_name} 中的列名未包含全部所需列，请检查文件列名设置！")
                custom_warning("提示", "shipment文件的列名不符合要求，请确认选择的文件是否正确！")
                raise ValueError("列名不符合要求")
                
        except FileNotFoundError:
            print(f"找不到文件 {shipment_file_name}，请确认文件路径是否正确。")
        except Exception as e:
            print(f"读取 {shipment_file_name} 文件出现其他错误: {e}")
            # 异常处理，隐藏进度条
            progress_bar.place_forget()
            # 清空文件选择相关显示信息示例（可根据实际需求调整）
            odm_gr_label.config(text="")
            suppllier_gr_label.config(text="")
            shipment_label.config(text="")
            result_label.config(text="数据处理出现异常，请检查文件或联系管理员！")
            result_label.update_idletasks()

        # TNB_KB_shipment_original = pd.read_excel(shipment_file_name, sheet_name='Summary', engine='openpyxl', usecols=columns_to_read)
        # 先移除'Share'列数据中的'%'符号，再转换为float类型
        # TNB_KB_shipment['Share'] = TNB_KB_shipment['Share'].str.rstrip('%').astype(float) / 100  # 除以100是将百分比转换为对应的小数形式，比如'30.00%'会变成0.3
        TNB_KB_shipment_original['Share'] = TNB_KB_shipment_original['Share'].astype(float)
        # print(TNB_KB_shipment['Share'])

        # 初始化写入excel的数据集对象
        TNB_KB_shipment_melted_excel = pd.DataFrame()  # 创建一个明细空的数据框
        TNB_KB_shipment_GroupBySuppliers_melted_excel = pd.DataFrame()  # 创建一个汇总空的数据框
        TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel = pd.DataFrame()  # 创建一个空实际比例饼图的数据框
        TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel = pd.DataFrame()  # 创建一个空理论比例饼图的数据框

        # 创建一个DataFrame，有"季度", "物料类型", "Suppliers"初始化的三例
        TNB_KB_shipment_GroupBySuppliers_original = TNB_KB_shipment_original.drop_duplicates(subset=["季度", "物料类型", "Suppliers"])

        # 创建一个DataFrame，有"Month"初始化的列
        GroupBySuppliers_original = Supplier_data.drop_duplicates(subset=["Month"])

        # 选取只需要保存的"季度", "物料类型", "Suppliers"这三列
        TNB_KB_shipment_GroupBySuppliers_original = TNB_KB_shipment_GroupBySuppliers_original[["季度", "物料类型", "Suppliers"]]

        # 对"季度"列去重并创建名为quarter的DataFrame
        quarter = pd.DataFrame({'季度': TNB_KB_shipment_GroupBySuppliers_original['季度'].drop_duplicates()})

        # 对"物料类型"列去重并创建名为materials的DataFrame
        materials = pd.DataFrame({'物料类型': TNB_KB_shipment_GroupBySuppliers_original['物料类型'].drop_duplicates()})

        #  对"年月"列去重并创建名为materials的DataFrame
        shipment_months = pd.DataFrame({'Month': GroupBySuppliers_original['Month'].drop_duplicates()})

        # 对'物料类型' "Suppliers"列去重并创建名为Suppliers的DataFrame
        Suppliers = pd.DataFrame({'Suppliers': TNB_KB_shipment_GroupBySuppliers_original['Suppliers'].drop_duplicates()})

        # 示例步骤一，完成后更新进度条
        # 此处省略你原代码里真实的数据处理逻辑，可参考原代码内容
        progress_bar["value"] += 20  # 假设这一步占总进度的20%，可根据实际调整
        root.update_idletasks()


        #先进行数据同类合并
        ODM_data = pd.DataFrame(ODM_data)
        Supplier_data = pd.DataFrame(Supplier_data)

        # 删除指定列ODM PN
        ODM_data = ODM_data.drop(labels='ODM PN', axis=1)
        Supplier_data = Supplier_data.drop(labels='ODM PN', axis=1)
        # Supplier_data = Supplier_data.drop(labels='标准命名', axis=1)

        #将Commodity列的数据统一替换为Keyboard
        ODM_data['Commodity'] = 'Keyboard'
        Supplier_data['Commodity'] = 'Keyboard'

        # sum = df.groupby(['ODM','PN']).sum()
        ODM_data = ODM_data.groupby(['BU','ODM','Lenovo PN','Vendor Name','Month','Commodity']).sum().reset_index()
        Supplier_data = Supplier_data.groupby(['BU','ODM','Lenovo PN','Vendor Name','Month','标准命名','Commodity']).sum().reset_index()

        #按行合并两个数据框
        #用于比较明细的表
        difference_data = pd.merge(ODM_data,Supplier_data,on=['BU','ODM','Lenovo PN','Vendor Name','Month','Commodity'], how='outer',suffixes=('_ODM数据值','_Supllier数据值'))
        #用于汇总数据的表

        # 用完后销毁ODM_data和Supplier_data
        del ODM_data
        del Supplier_data

        difference_data.fillna(0,inplace=True)

        #添加列标识差异数据来自于哪个表
        difference_data['Qty差值'] = difference_data['Qty_Supllier数据值'] - difference_data['Qty_ODM数据值']


        #重新定义列顺序
        new_column_order = ['BU','Month','ODM','Vendor Name','Lenovo PN','标准命名','Qty_Supllier数据值','Qty_ODM数据值','Qty差值','Commodity']
        difference_data = difference_data[new_column_order]

        #重新定义列名
        difference_data = difference_data.rename(columns={'Qty差值': 'Gap', 'Qty_Supllier数据值': 'Supllier QTY','Qty_ODM数据值':'ODM QTY'})

        #判断是否超过5000
        difference_data['边界值'] = difference_data['Gap'].apply(lambda x: "相等" if x == 0 else ("大于等于5000" if x >= 5000 else "小于5000"))

        # difference_data.to_excel('difference_data.xlsx')

        # 示例步骤二，完成后更新进度条
        # 更多原代码里的数据处理逻辑
        progress_bar["value"] += 30  # 假设这一步占总进度的30%，可根据实际调整
        root.update_idletasks()

        #------------------------------获取需要计算的季度和月份----------
        # 定义一个函数，将季度数据转换为对应的月份列表
        def quarter_to_months(quarter_str):
            year = int(quarter_str[:4])
            quarter_num = int(quarter_str[-1])
            start_month = (quarter_num - 1) * 3 + 1
            end_month = quarter_num * 3
            return [int(f"{year}{str(m).zfill(2)}") for m in range(start_month, end_month + 1)]

        # 将quarter中的季度数据转换为月份数据存到quarter_months中
        quarter_months = pd.DataFrame()
        for q in quarter['季度'].tolist():
            months_list = quarter_to_months(q)
            temp_df = pd.DataFrame({'Month': months_list})
            quarter_months = pd.concat([quarter_months, temp_df], ignore_index=True)

        # 取quarter_months和shipment_months的交集并更新quarter_months
        quarter_months = pd.merge(quarter_months, shipment_months, on='Month', how='inner')

        # 定义一个函数，将包含月份数据的DataFrame转换为指定格式的字典
        def months_df_to_dict(months_df):
            result_dict = {}
            for index, row in months_df.iterrows():
                month = row['Month']
                year = month // 10000
                month_num = month % 100
                quarter_num = (month_num - 1) // 3 + 1
                quarter = f"Q{quarter_num}"
                if quarter not in result_dict:
                    result_dict[quarter] = []
                result_dict[quarter].append(month)
            return result_dict

        # 调用函数进行转换
        quarter_months_dict = months_df_to_dict(quarter_months)

        #--------------------------------开始处理明细和汇总——————————————————————————————————————————————
        for index, row_q in quarter.iterrows():
            # 示例步骤三，完成后更新进度条
            # 继续原代码里的数据处理逻辑
            # progress_bar["value"] += 50  # 假设这一步占总进度的50%，可根据实际调整
            # root.update_idletasks()

            quarter_str = row_q['季度'][-2:]  # 获取季度标识中的季度部分，如从2024Q1中获取Q1
            year = int(row_q['季度'][:4])  # 从完整的季度标识字符串中提取前四位作为年份，修改此处获取年份的方式
            months = quarter_months_dict[quarter_str]
        
            #读取该季度数据
            TNB_KB_shipment = TNB_KB_shipment_original[TNB_KB_shipment_original['季度'] == row_q['季度']]
            # TNB_KB_shipment.to_excel('TNB_KB_shipment-3.xlsx')
            

            # 循环处理每个月份
            for month in months:
                
                #---------------------计算对应月份数据-----------------------------
                difference_data_original = difference_data[difference_data['Month'] == month]
                # ['{}实际产值'.format(quarter_str)]
                # difference_data_original.to_excel('{}2121.xlsx'.format(month))

                # 按照指定列对difference_data_original的Supllier QTY列进行汇总求和

                grouped = difference_data_original.groupby(['BU', 'Month', 'ODM', 'Vendor Name', '标准命名'])['Supllier QTY'].sum().reset_index()
                # grouped.to_excel('{}-group.xlsx'.format(month))
                # 根据条件匹配，将汇总后的值插入到TNB_KB_shipment对应的位置
                for index, row in grouped.iterrows():
                    matching_index_shipment = TNB_KB_shipment[
                        (TNB_KB_shipment['ODM'] == row['ODM'])
                        & (TNB_KB_shipment['Suppliers'] == row['Vendor Name'])
                        & (TNB_KB_shipment['KB Spec'] == row['标准命名'])
                        & (TNB_KB_shipment['季度'] == row_q['季度'])
                    ].index
                    if not matching_index_shipment.empty:
                        col_name = str(month).zfill(4)  # 构建对应列名，如202407
                        TNB_KB_shipment.loc[matching_index_shipment[0], col_name] = row['Supllier QTY']

                del difference_data_original
                del grouped
                # TNB_KB_shipment.to_excel('{}-kb.xlsx'.format(month))
            # TNB_KB_shipment.to_excel('TNB_KB_shipment-2.xlsx')
            # 值为空的补0 对应季度的各个月份
            for month in months:
                col_name = str(month).zfill(4)
                if col_name in TNB_KB_shipment.columns:
                    TNB_KB_shipment[col_name] = TNB_KB_shipment[col_name].fillna(0)
                else:
                    break

            # 季度实际产值
            quarter_actual_value_cols = [str(m).zfill(4) for m in months]

            # TNB_KB_shipment.to_excel('TNB_KB_shipment-1.xlsx')
            TNB_KB_shipment['{}实际产值'.format(quarter_str)] = TNB_KB_shipment[quarter_actual_value_cols].sum(axis=1)

            # TNB_KB_shipment.to_excel('TNB_KB_shipment.xlsx')

            #---------------------按型号汇总（对应季度总数）-----------------------------
            difference_data_originalTTL = difference_data[difference_data['Month'].isin(months)]

            groupedTTL = difference_data_originalTTL.groupby(['ODM', '标准命名'])['Supllier QTY'].sum().reset_index()

            # 根据条件匹配，将汇总后的值插入到TNB_KB_shipment对应的位置
            for index, row in groupedTTL.iterrows():
                matching_index_shipment = TNB_KB_shipment[
                    (TNB_KB_shipment['ODM'] == row['ODM'])
                    & (TNB_KB_shipment['KB Spec'] == row['标准命名'])
                    & (TNB_KB_shipment['季度'] == row_q['季度'])
                ].index
                

                if not matching_index_shipment.empty:
                    # 遍历所有符合条件的索引，为对应的行更新 'TTL' 列的值
                    for idx in matching_index_shipment:
                        TNB_KB_shipment.loc[idx, '{}总数'.format(quarter_str)] = row['Supllier QTY']

            del difference_data_originalTTL
            del groupedTTL

            # 对col1列的缺失值填充为0
            TNB_KB_shipment['{}总数'.format(quarter_str)] = TNB_KB_shipment['{}总数'.format(quarter_str)].fillna(0)
            # TNB_KB_shipment.to_excel('112.xlsx', index=False)

            # TNB_KB_shipment['{}理论产值'.format(quarter_str)] = TNB_KB_shipment['{}总数'.format(quarter_str)] * TNB_KB_shipment['Share']
            TNB_KB_shipment['{}理论产值'.format(quarter_str)] = TNB_KB_shipment['{}总数'.format(quarter_str)] * TNB_KB_shipment['Share']
            # 使用numpy.round函数对计算得到的每一个元素进行四舍五入保留整数操作
            TNB_KB_shipment['{}理论产值'.format(quarter_str)] = np.round(TNB_KB_shipment['{}理论产值'.format(quarter_str)].astype(float))

            # # 将Share列的数据转换为百分比格式
            TNB_KB_shipment['Share'] = TNB_KB_shipment['Share'].apply(lambda x: "{:.2%}".format(x))

            # 修改列名Share列名为根据季度动态显示列名
            TNB_KB_shipment.rename(columns={'Share': '{}Share'.format(quarter_str)}, inplace=True)
            

            # 算出Gap
            TNB_KB_shipment['{}Gap'.format(quarter_str)] = TNB_KB_shipment['{}理论产值'.format(quarter_str)] - TNB_KB_shipment['{}实际产值'.format(quarter_str)]

            # TNB_KB_shipment.to_excel('11.xlsx')

            # 计算并输出供应商的汇总信息-----------------------------------------------------
            TNB_KB_shipment_GroupBySuppliersTTL = TNB_KB_shipment

            # TNB_KB_shipment_GroupBySuppliersTTL.to_excel('122222.xlsx')

            # 读取该季度数据
            TNB_KB_shipment_GroupBySuppliers = TNB_KB_shipment_GroupBySuppliers_original[TNB_KB_shipment_GroupBySuppliers_original['季度'] == row_q['季度']]

            # 循环处理按供应商汇总每个月份数据
            for month in months:
                col_name =str(month).zfill(4)
                # 按照指定列对TNB_KB_shipment_GroupBySuppliersTTL的对应月份列进行汇总求和
                grouped_GroupBySuppliers = TNB_KB_shipment_GroupBySuppliersTTL.groupby(['季度','物料类型','Suppliers'])[col_name].sum().reset_index()
            
                
                # grouped = difference data original.groupby(['Bu', 'Month', 'oDM''vendor Name','标准命名'])['supllier Qry'].sum().reset index()

                # 根据条件匹配，将汇总后的值插入到TNB_KB_shipment对应的位置
                for index, row in grouped_GroupBySuppliers.iterrows():
                    matching_index_GroupBySuppliers = TNB_KB_shipment_GroupBySuppliers[
                        (TNB_KB_shipment_GroupBySuppliers['Suppliers'] == row['Suppliers'])
                        & (TNB_KB_shipment_GroupBySuppliers['物料类型'] == row['物料类型'])
                        & (TNB_KB_shipment_GroupBySuppliers['季度'] == row_q['季度'])
                    ].index
                    if not matching_index_GroupBySuppliers.empty:
                        supplier_col_name = '{}'.format(col_name)
                        TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliers[0], supplier_col_name] = row[col_name]
                        TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliers[0], col_name] = TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliers[0], col_name].astype(int)

                del grouped_GroupBySuppliers

            # 按Q3汇总供应商实际产值
            grouped_GroupBySuppliersActual = TNB_KB_shipment_GroupBySuppliersTTL.groupby(['季度','物料类型','Suppliers'])['{}实际产值'.format(quarter_str)].sum().reset_index()
            for index, row in grouped_GroupBySuppliersActual.iterrows():
                matching_index_GroupBySuppliersActual = TNB_KB_shipment_GroupBySuppliers[
                    (TNB_KB_shipment_GroupBySuppliers['Suppliers'] == row['Suppliers'])
                    & (TNB_KB_shipment_GroupBySuppliers['物料类型'] == row['物料类型'])
                    & (TNB_KB_shipment_GroupBySuppliers['季度'] == row_q['季度'])
                ].index
                if not matching_index_GroupBySuppliersActual.empty:
                    TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersActual[0], '{}实际总值'.format(quarter_str)] = row['{}实际产值'.format(quarter_str)]
                    #将整数值列格式化成int型
                    TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersActual[0], col_name] = TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersActual[0], col_name].astype(int)
            del grouped_GroupBySuppliersActual

            # 按Q3汇总供应商理论产值
            grouped_GroupBySuppliersTheory = TNB_KB_shipment_GroupBySuppliersTTL.groupby(['季度','物料类型','Suppliers'])['{}理论产值'.format(quarter_str)].sum().reset_index()
            for index, row in grouped_GroupBySuppliersTheory.iterrows():
                matching_index_GroupBySuppliersTheory = TNB_KB_shipment_GroupBySuppliers[
                    (TNB_KB_shipment_GroupBySuppliers['Suppliers'] == row['Suppliers'])
                    & (TNB_KB_shipment_GroupBySuppliers['物料类型'] == row['物料类型'])
                    & (TNB_KB_shipment_GroupBySuppliers['季度'] == row_q['季度'])
                ].index
                if not matching_index_GroupBySuppliersTheory.empty:
                    TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersTheory[0], '{}理论总值'.format(quarter_str)] = row['{}理论产值'.format(quarter_str)]
                    #将整数值列格式化成int型
                    TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersTheory[0], col_name] = TNB_KB_shipment_GroupBySuppliers.loc[matching_index_GroupBySuppliersTheory[0], col_name].astype(int)
        
            del grouped_GroupBySuppliersTheory
            del TNB_KB_shipment_GroupBySuppliersTTL

            # 示例步骤三，完成后更新进度条
            # 继续原代码里的数据处理逻辑
            progress_bar["value"] += 90  # 假设这一步占总进度的50%，可根据实际调整
            root.update_idletasks()


            # 算出供应商对应季度实际生产的比例数----------------------------------------------------------
            total_Q3_actual_value = TNB_KB_shipment_GroupBySuppliers['{}实际总值'.format(quarter_str)].sum()
            TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)] = TNB_KB_shipment_GroupBySuppliers['{}实际总值'.format(quarter_str)].apply(lambda x: x / total_Q3_actual_value)
            # TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)] = TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)].map(lambda x: "{:.2%}".format(x))
            TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)] = TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)].astype(float)
            

            # 算出供应商对应季度理论生产的比例数----------------------------------------------------------
            total_Q3_theory_value = TNB_KB_shipment_GroupBySuppliers['{}理论总值'.format(quarter_str)].sum()
            TNB_KB_shipment_GroupBySuppliers['Share'] = TNB_KB_shipment_GroupBySuppliers['{}理论总值'.format(quarter_str)].apply(lambda x: x / total_Q3_theory_value)
            # TNB_KB_shipment_GroupBySuppliers['Share'] = TNB_KB_shipment_GroupBySuppliers['Share'].map(lambda x: "{:.2%}".format(x))
            TNB_KB_shipment_GroupBySuppliers['Share'] = TNB_KB_shipment_GroupBySuppliers['Share'].astype(float)

            # 修改列名Share列名为根据季度动态显示列名
            TNB_KB_shipment_GroupBySuppliers.rename(columns={'Share': '{}Share'.format(quarter_str)}, inplace=True)

            # 算出数量Gap
            TNB_KB_shipment_GroupBySuppliers['{}Gap_数量'.format(quarter_str)] = TNB_KB_shipment_GroupBySuppliers['{}理论总值'.format(quarter_str)] - TNB_KB_shipment_GroupBySuppliers['{}实际总值'.format(quarter_str)]

            # 算出Share的Gap
            TNB_KB_shipment_GroupBySuppliers['{}Gap_共享'.format(quarter_str)] = TNB_KB_shipment_GroupBySuppliers['{}Share'.format(quarter_str)] - TNB_KB_shipment_GroupBySuppliers['{}实际生产比例'.format(quarter_str)]

            # TNB_KB_shipment_GroupBySuppliers.to_excel('21.xlsx')

            #将明细进行逆透视列
            # 定义不参与逆透视的列，也就是作为标识符的列
            id_columns = ['季度', '物料类型', 'ODM', 'KB Spec', 'Suppliers']
            # 执行逆透视操作
            TNB_KB_shipment_melted = TNB_KB_shipment.melt(
                id_vars=id_columns,
                var_name='指标',
                value_name='数据值'
            )

            # 使用add方法将更新后的数据累加到TNB_KB_shipment_melted_excel中，fill_value=0处理可能的缺失值情况
            TNB_KB_shipment_melted_excel = pd.concat([TNB_KB_shipment_melted_excel, TNB_KB_shipment_melted], axis=0, ignore_index=True) 
            
            # TNB_KB_shipment_melted_excel = TNB_KB_shipment_melted_excel.add(TNB_KB_shipment_melted, fill_value=0)

            #将汇总进行逆透视列
            # 定义不参与逆透视的列，也就是作为标识符的列
            id_columns = ['季度', '物料类型', 'Suppliers']
            # 执行逆透视操作
            TNB_KB_shipment_GroupBySuppliers_melted = TNB_KB_shipment_GroupBySuppliers.melt(
                id_vars=id_columns,
                var_name='指标',
                value_name='数据值'
            )

            # 使用add方法将更新后的数据累加到TNB_KB_shipment_GroupBySuppliers_melted中，fill_value=0处理可能的缺失值情况
            TNB_KB_shipment_GroupBySuppliers_melted_excel = pd.concat([TNB_KB_shipment_GroupBySuppliers_melted_excel, TNB_KB_shipment_GroupBySuppliers_melted], axis=0, ignore_index=True) 
            # TNB_KB_shipment_GroupBySuppliers_melted_excel = TNB_KB_shipment_GroupBySuppliers_melted_excel.add(TNB_KB_shipment_GroupBySuppliers_melted, fill_value=0)

            

            #输出用于显示饼图的数据---供应商实际生产总值----------------
            # 筛选出需要的列，包含名字中含'Q'的列以及'季度'、'物料类型'、'Suppliers'列
            columns_to_keep = ['季度', '物料类型', 'Suppliers']
            columns_to_keep.extend([col for col in TNB_KB_shipment_GroupBySuppliers.columns if '实际总值' in col])

            # 创建新的数据集TNB_KB_shipment_GroupBySuppliers_pie_char
            TNB_KB_shipment_GroupBySuppliers_pie_actual = TNB_KB_shipment_GroupBySuppliers[columns_to_keep].copy()

            #将汇总的饼图进行逆透视列 ----供应商实际生产总值
            # 定义不参与逆透视的列，也就是作为标识符的列
            id_columns = ['季度', '物料类型', 'Suppliers']
            # 执行逆透视操作
            TNB_KB_shipment_GroupBySuppliers_pie_actual_melted = TNB_KB_shipment_GroupBySuppliers_pie_actual.melt(
                id_vars=id_columns,
                var_name='指标',
                value_name='数据值'
            )

            # 使用add方法将更新后的数据累加到TTNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel中，fill_value=0处理可能的缺失值情况
            TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel = pd.concat([TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel, TNB_KB_shipment_GroupBySuppliers_pie_actual_melted], axis=0, ignore_index=True) 
            # TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel = TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel.add(TNB_KB_shipment_GroupBySuppliers_pie_actual_melted, fill_value=0)


            #输出用于显示饼图的数据---供应商理论生产总值----------------
            # 筛选出需要的列，包含名字中含'Q'的列以及'季度'、'物料类型'、'Suppliers'列
            columns_to_keep = ['季度', '物料类型', 'Suppliers']
            columns_to_keep.extend([col for col in TNB_KB_shipment_GroupBySuppliers.columns if '理论总值' in col])

            # 创建新的数据集TNB_KB_shipment_GroupBySuppliers_pie_char
            TNB_KB_shipment_GroupBySuppliers_pie_theory = TNB_KB_shipment_GroupBySuppliers[columns_to_keep].copy()

            #将汇总的饼图进行逆透视列 ----供应商理论生产总值
            # 定义不参与逆透视的列，也就是作为标识符的列
            id_columns = ['季度', '物料类型', 'Suppliers']
            # 执行逆透视操作
            TNB_KB_shipment_GroupBySuppliers_pie_theory_melted = TNB_KB_shipment_GroupBySuppliers_pie_theory.melt(
                id_vars=id_columns,
                var_name='指标',
                value_name='数据值'
            )

            # 使用add方法将更新后的数据累加到TTNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel中，fill_value=0处理可能的缺失值情况
            TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel = pd.concat([TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel, TNB_KB_shipment_GroupBySuppliers_pie_theory_melted], axis=0, ignore_index=True) 
            
            # TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel = TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel.add(TNB_KB_shipment_GroupBySuppliers_pie_theory_melted, fill_value=0)

        with pd.ExcelWriter('shipment.xlsx') as writer:
            #将GR Analysis明细写到excel中
            difference_data.to_excel(writer, sheet_name='GR Analysis', index=False)
            
            # 将TNB_KB_shipment的数据写入到名为"明细"的sheet页
            TNB_KB_shipment_melted_excel.to_excel(writer, sheet_name='明细', index=False)

            # 将TNB_KB_shipment_GroupBySuppliers_melted的数据写入到名为"汇总"的sheet页
            TNB_KB_shipment_GroupBySuppliers_melted_excel.to_excel(writer, sheet_name='汇总', index=False)

            # 将TNB_KB_shipment_GroupBySuppliers的数据写入到名为"汇总"的sheet页----用于显示供应商实际生产饼图
            TNB_KB_shipment_GroupBySuppliers_pie_actual_melted_excel.to_excel(writer, sheet_name='供应商实际生产汇总饼图', index=False)

            # 将TNB_KB_shipment_GroupBySuppliers的数据写入到名为"汇总"的sheet页----用于显示供应商理论生产饼图
            TNB_KB_shipment_GroupBySuppliers_pie_theory_melted_excel.to_excel(writer, sheet_name='供应商理论生产汇总饼图', index=False)

            # quarter"的数据写入sheet页
            quarter.to_excel(writer, sheet_name='quarter', index=False)

            # materials"的数据写入sheet页
            materials.to_excel(writer, sheet_name='materials', index=False)

            # Suppliers"的数据写入sheet页
            Suppliers.to_excel(writer, sheet_name='Suppliers', index=False)


        # 加载Excel文件以获取工作簿对象
        # wb = load_workbook(file_path)
        
        current_path = os.getcwd()
        file_path = current_path + '\shipment.xlsx'
        result_label.config(text="清洗后的文件位置: " + file_path + "；请到Power BI中刷新和分析")
        result_label.update_idletasks()  # 强制更新界面显示文件路径
    except Exception as e:
        print(f"执行 long_running_operation 函数出现错误: {e}")
        # 异常处理，隐藏进度条
        progress_bar.place_forget()
        # 清空文件选择相关显示信息示例（可根据实际需求调整）
        odm_gr_label.config(text="")
        suppllier_gr_label.config(text="")
        shipment_label.config(text="")
        result_label.config(text="数据处理出现异常，请检查文件或联系管理员！")
        result_label.update_idletasks()


def on_cancel():
    root.destroy()


root = tk.Tk()
root.title("数据清洗")
root.geometry("750x380")
root.resizable(False, False)

# 设置图标
root.iconbitmap("数据清洗.ico")

# 让窗口先更新下，确保能获取到准确的窗口尺寸信息
root.update_idletasks()

# 计算窗口在屏幕中心的坐标
x = (root.winfo_screenwidth() - root.winfo_width()) // 2
y = (root.winfo_screenheight() - root.winfo_height()) // 2
root.geometry(f"+{x}+{y}")

# 用于存储选择的文件路径
odm_gr_file_name = ""
suppllier_gr_file_name = ""
shipment_file_name = ""

# 选择 ODM GR.xlsx 文件的按钮及相关逻辑
file1_button = tk.Button(root, text="点此选择ODM GR数据", command=lambda: get_file_name('ODM GR'), width=24, height=2, font=("TkDefaultFont", 12), anchor=tk.W)
file1_button.pack(anchor=tk.W)
odm_gr_label = tk.Label(root, text="", fg="blue")
odm_gr_label.pack(anchor=tk.W)


# 选择 Supplier GR.xlsx 文件的按钮及相关逻辑
file2_button = tk.Button(root, text="点此选择Supplier GR数据", command=lambda: get_file_name('Supplier GR'), width=24, height=2, font=("TkDefaultFont", 12), anchor=tk.W)
file2_button.pack(anchor=tk.W)
suppllier_gr_label = tk.Label(root, text="", fg="blue")
suppllier_gr_label.pack(anchor=tk.W)


# 选择shipment.xlsx 文件的按钮及相关逻辑
file3_button = tk.Button(root, text="点此选择shipment数据", command=lambda: get_file_name('TNB KB shipment'), width=24, height=2, font=("TkDefaultFont", 12), anchor=tk.W)
file3_button.pack(anchor=tk.W)
shipment_label = tk.Label(root, text="", fg="blue")
shipment_label.pack(anchor=tk.W)

# 创建一个外层框架，用于调整按钮整体的垂直位置
outer_frame = tk.Frame(root)
outer_frame.pack(fill=tk.BOTH, expand=True)

# 创建一个内层框架用于放置按钮，采用grid布局方式，并让其靠窗口底部摆放
button_frame = tk.Frame(outer_frame)
button_frame.pack(side=tk.BOTTOM, fill=tk.X)

# 调整按钮位置，使用grid布局将按钮放置在中间，设置合适间距和对齐方式让按钮在水平和垂直方向布局合理
proceed_button = tk.Button(button_frame, text="确定", command=proceed, width=10, height=2, font=("TkDefaultFont", 12), fg="blue")
cancel_button = tk.Button(button_frame, text="关闭", command=on_cancel,width=10, height=2, font=("TkDefaultFont", 12), fg="blue")
proceed_button.grid(row=0, column=0, padx=(10, 30), pady=10, sticky="e")
cancel_button.grid(row=0, column=1, padx=(30, 10), pady=10, sticky="w")

# 设置button_frame的列权重，让两列均匀分配空间，使得按钮能在水平方向居中
button_frame.columnconfigure(0, weight=1)
button_frame.columnconfigure(1, weight=1)

result_label = tk.Label(root, text="",fg='red')
result_label.pack(anchor='center', padx=40, pady=40)

# 创建进度条组件
progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=750, mode='determinate')
progress_bar.place(relx=0.25, rely=0.85, relwidth=0.5, height=30)
#隐藏进度条（也可以选择不隐藏，根据实际需求）
progress_bar.place_forget()

#自定义弹出框
def custom_warning(title, message):
    top = tk.Toplevel()
    top.title(title)
    top.geometry("300x100")
    top.resizable(False, False)
    # 设置弹出框的图标，将这里的'your_icon.ico'替换为你实际的图标文件路径及文件名
    top.iconbitmap('数据清洗.ico')

    # 获取父窗口（root）相对于屏幕左上角的绝对坐标以及尺寸信息
    root_x = root.winfo_rootx()
    root_y = root.winfo_rooty()
    root_width = root.winfo_width()
    root_height = root.winfo_height()

    # 计算弹出框在父窗口中心的位置
    top_width = 300
    top_height = 100
    x = root_x + (root_width - top_width) // 2
    y = root_y + (root_height - top_height) // 2
    top.geometry("+{}+{}".format(x, y))

    # 设置字体及其他样式（可进一步优化，比如添加背景色等）
    custom_font = tkFont.Font(family="TkDefaultFont", size=8)  # 增大字号示例
    label = tk.Label(top, text=message, font=custom_font, fg="red")  # 添加白色背景色示例
    label.pack(pady=20)

    button = tk.Button(top, text="确定", command=top.destroy, width=8, height=2)
    button.pack()


# 通用的文件选择函数，根据传入的文件名提示选择对应文件
def get_file_name(file_type):
    global odm_gr_file_name, suppllier_gr_file_name, shipment_file_name
    if file_type == 'ODM GR':
        file_name = filedialog.askopenfilename(title="选择ODM GR.xlsx文件", filetypes=[("Excel files", "*.xlsx")])
        if file_name:
            odm_gr_file_name = file_name
            odm_gr_label.config(text="ODM GR数据: " + file_name)
    elif file_type == 'Supplier GR':
        file_name = filedialog.askopenfilename(title="选择Supplier GR.xlsx文件", filetypes=[("Excel files", "*.xlsx")])
        if file_name:
            suppllier_gr_file_name = file_name
            suppllier_gr_label.config(text="Supplier GR数据: " + file_name)
    elif file_type == 'TNB KB shipment':
        file_name = filedialog.askopenfilename(title="选择TNB KB shipment.xlsx文件", filetypes=[("Excel files", "*.xlsx")])
        if file_name:
            shipment_file_name = file_name
            shipment_label.config(text="TNB KB shipment数据: " + file_name)


root.mainloop()